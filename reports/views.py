from django.shortcuts import render
from companies.models import Company
from jobs.models import Job
from resumes.models import Resume
from users.models import User
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series, LineChart, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from django.http import HttpResponse
from django.db.models import Count
from openpyxl.styles import NamedStyle
from openpyxl.chart.label import DataLabelList


class ReportView(APIView):
    def get(self, request, *args, **kwargs):
        one_week_ago = datetime.now() - timedelta(days=7)
        company_count = Company.objects.filter(created_at__gte=one_week_ago).count()
        job_count = Job.objects.filter(created_at__gte=one_week_ago).count()
        resume_count = Resume.objects.filter(created_at__gte=one_week_ago).count()
        user_count = User.objects.filter(created_at__gte=one_week_ago).count()

        return Response({
            'company_count': company_count,
            'job_count': job_count,
            'resume_count': resume_count,
            'user_count': user_count
        }, status=status.HTTP_200_OK)

class ExcelReportView(APIView):
    def get(self, request, *args, **kwargs):
        # Lấy dữ liệu
        companies = Company.objects.all()
        jobs = Job.objects.all()
        resumes = Resume.objects.all()
        users = User.objects.all()

        # Tạo Workbook
        wb = Workbook()

        # Tạo sheet cho Users
        ws_users = wb.active
        ws_users.title = "Users"
        user_data = [{'ID': user.id, 'Tên người dùng': user.name, 'Vai trò': user.role.name, 'Số lượng hồ sơ đã nộp': 0} for user in users]
        df_users = pd.DataFrame(user_data)

        # Tạo biểu đồ cho Users
        user_resume_counts = resumes.values('user').annotate(count=Count('id'))
        user_resume_data = [{'user': User.objects.get(id=item['user']).name, 'resume_count': item['count']} for item in user_resume_counts]
        df_user_resume = pd.DataFrame(user_resume_data)

        # Cập nhật số lượng hồ sơ đã nộp vào df_users
        for index, row in df_users.iterrows():
            user_name = row['Tên người dùng']
            if user_name in df_user_resume['user'].values:
                resume_count = df_user_resume[df_user_resume['user'] == user_name]['resume_count'].values[0]
                df_users.at[index, 'Số lượng hồ sơ đã nộp'] = resume_count

        # Thêm dữ liệu vào sheet Users
        for r in dataframe_to_rows(df_users, index=False, header=True):
            ws_users.append(r)

        # Tự động điều chỉnh độ rộng của các cột
        for col in ws_users.columns:
            max_length = 0
            column = col[0].column_letter  # Lấy tên cột
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_users.column_dimensions[column].width = adjusted_width

        chart = BarChart()
        data = Reference(ws_users, min_col=4, min_row=1, max_col=4, max_row=len(df_users) + 1)
        categories = Reference(ws_users, min_col=2, min_row=2, max_row=len(df_users) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = "Hồ sơ theo người dùng"
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showPercent = False
        chart.dLbls.showCatName = False
        chart.dLbls.showSerName = False
        chart.dLbls.showLegendKey = False

        ws_users.add_chart(chart, "E5")

        # Tạo sheet cho Companies
        ws_companies = wb.create_sheet(title="Companies")
        company_data = [{'ID': company.id, 'Tên công ty': company.name, 'Địa chỉ': company.address, 'Ngày tạo': company.created_at.replace(tzinfo=None), 'Số lượng công việc đang tuyển': Job.objects.filter(company=company).count(), 'Số lượng người theo dõi': company.users_followed.count(), 'Số lượng người đã nộp hồ sơ': resumes.filter(job__company=company).count()} for company in companies]
        df_companies = pd.DataFrame(company_data)
        for r in dataframe_to_rows(df_companies, index=False, header=True):
            ws_companies.append(r)

        # Định dạng cột ngày tháng
        date_style = NamedStyle(name="date_style", number_format="DD-MM-YYYY")
        for cell in ws_companies['D']:
            cell.style = date_style

        # Tự động điều chỉnh độ rộng của các cột
        for col in ws_companies.columns:
            max_length = 0
            column = col[0].column_letter  # Lấy tên cột
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_companies.column_dimensions[column].width = adjusted_width

        # Tạo biểu đồ Bar cho số lượng người theo dõi
        bar_chart = BarChart()
        data = Reference(ws_companies, min_col=6, min_row=1, max_col=6, max_row=len(df_companies) + 1)
        categories = Reference(ws_companies, min_col=2, min_row=2, max_row=len(df_companies) + 1)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        bar_chart.title = "Số lượng người theo dõi theo công ty"
        bar_chart.dLbls = DataLabelList()
        bar_chart.dLbls.showVal = False
        bar_chart.dLbls.showPercent = False
        bar_chart.dLbls.showCatName = False
        bar_chart.dLbls.showSerName = False
        bar_chart.dLbls.showLegendKey = False
        ws_companies.add_chart(bar_chart, "E5")

        # Tạo biểu đồ Bar Group cho số lượng công việc đang tuyển và số lượng người đã nộp hồ sơ
        bar_group_chart = BarChart()
        data_jobs = Reference(ws_companies, min_col=5, min_row=1, max_col=5, max_row=len(df_companies) + 1)
        data_resumes = Reference(ws_companies, min_col=7, min_row=1, max_col=7, max_row=len(df_companies) + 1)
        bar_group_chart.add_data(data_jobs, titles_from_data=True)
        bar_group_chart.add_data(data_resumes, titles_from_data=True)
        bar_group_chart.set_categories(categories)
        bar_group_chart.title = "Công việc và hồ sơ theo công ty"
        bar_group_chart.style = 12  # Đặt kiểu biểu đồ
        bar_group_chart.grouping = "clustered"  # Đặt kiểu nhóm
        bar_group_chart.overlap = 0  # Đặt độ chồng chéo giữa các thanh
    
        # Đặt màu sắc cho các thanh
        for i, series in enumerate(bar_group_chart.series):
            if i == 0:
                series.graphicalProperties.solidFill = "4F81BD"  # Màu xanh
            elif i == 1:
                series.graphicalProperties.solidFill = "C0504D"  # Màu đỏ

        bar_group_chart.dLbls = DataLabelList()
        bar_group_chart.dLbls.showVal = False
        bar_group_chart.dLbls.showPercent = False
        bar_group_chart.dLbls.showCatName = False
        bar_group_chart.dLbls.showSerName = False
        bar_group_chart.dLbls.showLegendKey = False

        ws_companies.add_chart(bar_group_chart, "E20")

        # Tạo sheet cho Jobs
        ws_jobs = wb.create_sheet(title="Jobs")
        job_data = [{'ID': job.id, 'Tên công việc': job.name, 'Lương': job.salary, 'Trình độ': job.level, 'Thuộc công ty': job.company.name} for job in jobs]
        df_jobs = pd.DataFrame(job_data)
        for r in dataframe_to_rows(df_jobs, index=False, header=True):
            ws_jobs.append(r)

        # Tự động điều chỉnh độ rộng của các cột
        for col in ws_jobs.columns:
            max_length = 0
            column = col[0].column_letter  # Lấy tên cột
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_jobs.column_dimensions[column].width = adjusted_width

        # Tạo biểu đồ Bar cho Jobs

        bar_chart = BarChart()
        data = Reference(ws_jobs, min_col=3, min_row=1, max_col=3, max_row=len(job_data) + 1)
        categories = Reference(ws_jobs, min_col=2, min_row=2, max_row=len(job_data) + 1)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        bar_chart.title = "Công việc theo mức lương"
        bar_chart.dLbls = DataLabelList()
        bar_chart.dLbls.showVal = True
        bar_chart.dLbls.showPercent = False
        bar_chart.dLbls.showCatName = False
        bar_chart.dLbls.showSerName = False
        bar_chart.dLbls.showLegendKey = False

        ws_jobs.add_chart(bar_chart, "E5")

        # Tạo biểu đồ Pie cho mức lương của các công việc
        pie_chart = PieChart()
        pie_data = Reference(ws_jobs, min_col=3, min_row=2, max_row=len(job_data) + 1)
        pie_categories = Reference(ws_jobs, min_col=2, min_row=2, max_row=len(job_data) + 1)
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_categories)
        pie_chart.title = "Công việc theo mức lương"
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showVal = False
        pie_chart.dataLabels.showCatName = False
        pie_chart.dataLabels.showSerName = False
        pie_chart.dataLabels.showLegendKey = False
        ws_jobs.add_chart(pie_chart, "E20")

        # Thiết lập header cho response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Thống-Kê.xlsx'

        # Ghi Workbook vào response
        wb.save(response)
        return response
